#!/usr/bin/env python3
"""
우주시스템 견적서 엑셀의 "부품단가" 시트를 Supabase parts 테이블로 임포트한다.

엑셀 구조:
  A열에 카테고리명이 있는 행이 블록의 시작이자 그 블록의 헤더 행이다.
  같은 행의 B~M열에 컬럼명(고유번호/제품명/가격/품절여부/…/등급/링크)이 들어 있고,
  다음 카테고리 행이 나오기 전까지가 그 카테고리의 품목이다.
  카테고리마다 스펙 컬럼이 다르므로 위치를 하드코딩하지 않고 헤더명으로 매핑한다.

재실행 안전: (category, name) 자연키로 upsert 한다.

사용법:
  python3 scripts/import_parts.py <엑셀경로> [--dry-run]

환경변수는 apps/web/.env.local에서 읽는다.
"""

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl이 필요합니다: pip3 install openpyxl")

REPO_ROOT = Path(__file__).resolve().parent.parent
ENV_FILE = REPO_ROOT / "apps" / "web" / ".env.local"
SHEET_NAME = "부품단가"

# 엑셀 A열의 카테고리명 → (DB enum, platform)
CATEGORY_MAP = {
    "AMD": ("cpu", "amd"),
    "AMD 마더보드": ("mainboard", "amd"),
    "INTEL": ("cpu", "intel"),
    "INTEL 마더보드": ("mainboard", "intel"),
    "메모리": ("memory", "common"),
    "SSD": ("ssd", "common"),
    "HDD": ("hdd", "common"),
    "그래픽카드": ("gpu", "common"),
    "파워": ("psu", "common"),
    "케이스": ("case", "common"),
    "CPU쿨러": ("cpu_cooler", "common"),
    "케이스팬": ("case_fan", "common"),
    "RGB 컨트롤러": ("rgb_controller", "common"),
    "공임 & AS": ("labor_as", "common"),
    "키보드": ("keyboard", "common"),
    "마우스": ("mouse", "common"),
    "스피커": ("speaker", "common"),
    "헤드셋": ("headset", "common"),
    "모니터": ("monitor", "common"),
    "SSD방열판": ("ssd_heatsink", "common"),
    "메모리방열판": ("memory_heatsink", "common"),
    "튜닝": ("tuning", "common"),
    "추가 품목": ("extra", "common"),
}

# 공통 컬럼 — specs에는 넣지 않는다
COMMON_HEADERS = {"고유번호", "제품명", "가격", "품절여부", "등급", "링크"}

# 스펙 헤더 → specs jsonb 키. 호환성 검증에 쓰이는 값은 이름을 명시적으로 고정한다.
SPEC_KEY_MAP = {
    "코어/쓰레드": "cores_threads",
    "코어클럭(GHz)": "clock_ghz",
    "캐시(MB)": "cache_mb",
    "메모리지원": "memory_support",   # 마더보드 DDR4/DDR5 — 메모리 호환성 판정
    "내장그래픽": "igpu",
    "타입": "form_factor",
    "전원부": "vrm",
    "용량(GB)": "capacity_gb",
    "종류(DDRx)": "ddr_type",         # 메모리 규격 — 마더보드와 대조
    "클럭(MHz)": "clock_mhz",
    "제원": "spec",
    "방식": "interface",
    "읽기속도(MB/s)": "read_mbps",
    "쓰기속도(MB/s)": "write_mbps",
    "메모리방식": "nand_type",
    "RPM": "rpm",
    "크기": "size",
    "가로길이(mm)": "length_mm",       # 그래픽카드 길이 — 케이스와 대조
    "부스트클럭(MHz)": "boost_mhz",
    "쿠다코어(개)": "cuda_cores",
    "메모리종류": "vram_type",
    "메모리클럭(MHz)": "vram_clock_mhz",
    "메모리용량(GB)": "vram_gb",
    "용량(W)": "watt",
    "80PLUS": "efficiency",
    "최대 +12V": "max_12v",
    "무상AS(년)": "warranty_years",
    "컬러": "color",
    "색상": "color",
    "수랭여부": "liquid_cooling",
    "GPU허용": "gpu_max_mm",           # 케이스가 허용하는 GPU 길이
    "CPU쿨러허용": "cooler_max_mm",     # 케이스가 허용하는 쿨러 높이
    # "높이(공랭)/열(수랭)"은 한 칸에 두 단위가 섞여 있어 expand_cooler_size()가 따로 처리한다.
    "TDP": "tdp",
    "크기(인치)": "size_inch",
    "해상도": "resolution",
    "주사율(Hz)": "refresh_hz",
    "패널": "panel",
}

# specs에서 숫자로 저장할 키 (호환성 계산에 쓰이므로 문자열이면 곤란하다)
NUMERIC_SPEC_KEYS = {
    "length_mm", "gpu_max_mm", "cooler_max_mm", "height_mm",
    "watt", "capacity_gb", "vram_gb", "cache_mb", "clock_mhz",
    "boost_mhz", "cuda_cores", "vram_clock_mhz", "rpm",
    "read_mbps", "write_mbps", "refresh_hz", "warranty_years",
}


def load_env() -> tuple[str, str]:
    if not ENV_FILE.exists():
        sys.exit(f"환경변수 파일을 찾을 수 없습니다: {ENV_FILE}")
    env = {}
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    url = env.get("NEXT_PUBLIC_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY 가 필요합니다")
    return url.rstrip("/"), key


def clean(value) -> str:
    """셀 값을 문자열로 정규화. 연속 공백을 하나로 줄인다."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def to_number(text: str):
    """'360mm', '1,390,000', '4.0~5.4' 같은 값에서 대표 숫자를 뽑는다."""
    if not text:
        return None
    # 범위 표기는 최댓값을 쓴다 (예: 코어클럭 3.5~4.4 → 4.4)
    nums = re.findall(r"\d+(?:[.,]\d+)?", text.replace(",", ""))
    if not nums:
        return None
    try:
        val = float(nums[-1] if "~" in text else nums[0])
    except ValueError:
        return None
    return int(val) if val.is_integer() else val


COOLER_SIZE_HEADER = "높이(공랭)/열(수랭)"


def expand_cooler_size(raw: str) -> dict:
    """
    CPU쿨러의 "높이(공랭)/열(수랭)" 한 칸에 두 가지 단위가 섞여 있다.
      "157"  → 공랭, 높이 157mm
      "3열"  → 수랭, 라디에이터 3열(=360mm)
    케이스의 CPU쿨러허용 높이와 비교하려면 둘을 구분해야 하므로 나눠 담는다.
    """
    if not raw:
        return {}
    if "열" in raw:
        rows = to_number(raw)
        if rows is None:
            return {}
        rows = int(rows)
        return {
            "cooler_type": "liquid",
            "radiator_rows": rows,
            # 1열 = 120mm 팬 1개 기준
            "radiator_mm": rows * 120,
        }
    height = to_number(raw)
    if height is None:
        return {}
    return {"cooler_type": "air", "height_mm": int(height)}


def parse_price(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^\d]", "", str(value))
    return int(digits) if digits else 0


def parse_sheet(path: Path) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"'{SHEET_NAME}' 시트가 없습니다. 있는 시트: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    parts: list[dict] = []
    warnings: list[str] = []
    category = platform = None
    headers: dict[int, str] = {}
    unknown_categories: set[str] = set()

    for row in range(1, ws.max_row + 1):
        label = clean(ws.cell(row, 1).value)

        # 카테고리 블록 시작 = 헤더 행
        if label:
            if label not in CATEGORY_MAP:
                unknown_categories.add(label)
                category = platform = None
                continue
            category, platform = CATEGORY_MAP[label]
            headers = {}
            for col in range(2, 20):
                name = clean(ws.cell(row, col).value)
                if name and name != "여백":
                    headers[col] = name
            continue

        if category is None:
            continue

        cells = {headers[c]: clean(ws.cell(row, c).value) for c in headers}
        name = cells.get("제품명", "")
        if not name or name == "0":
            continue

        specs = {}
        for header, raw in cells.items():
            if header in COMMON_HEADERS or not raw:
                continue
            if header == COOLER_SIZE_HEADER:
                specs.update(expand_cooler_size(raw))
                continue
            key = SPEC_KEY_MAP.get(header)
            if key is None:
                # 매핑에 없는 스펙은 원래 헤더명 그대로 보존한다 (유실 방지)
                key = header
            specs[key] = to_number(raw) if key in NUMERIC_SPEC_KEYS else raw

        part_no_raw = cells.get("고유번호", "")
        part_no = int(part_no_raw) if part_no_raw.isdigit() else None
        if part_no is None:
            warnings.append(f"고유번호 없음 → 자연키로 임포트: [{category}] {name}")

        parts.append({
            "part_no": part_no,
            "category": category,
            "platform": platform,
            "name": name,
            "price": parse_price(cells.get("가격")),
            "sold_out": cells.get("품절여부", "") == "품절",
            "grade": cells.get("등급") or None,
            "link": cells.get("링크") or None,
            "specs": specs,
            "active": True,
        })

    for cat in sorted(unknown_categories):
        warnings.append(f"매핑에 없는 카테고리라 건너뜀: {cat}")

    return parts, warnings


def upsert(url: str, key: str, parts: list[dict], batch: int = 100) -> int:
    endpoint = f"{url}/rest/v1/parts?on_conflict=category,name"
    done = 0
    for i in range(0, len(parts), batch):
        chunk = parts[i:i + batch]
        req = urllib.request.Request(
            endpoint,
            data=json.dumps(chunk, ensure_ascii=False).encode("utf-8"),
            headers={
                "apikey": key,
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=minimal",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req) as resp:
                resp.read()
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", "replace")
            sys.exit(f"\n임포트 실패 (행 {i}~{i+len(chunk)}): HTTP {e.code}\n{body}")
        done += len(chunk)
        print(f"  {done}/{len(parts)} …", end="\r", flush=True)
    print()
    return done


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="견적서 엑셀 파일 경로")
    ap.add_argument("--dry-run", action="store_true", help="DB에 쓰지 않고 파싱 결과만 출력")
    args = ap.parse_args()

    path = Path(args.xlsx).expanduser()
    if not path.exists():
        sys.exit(f"파일이 없습니다: {path}")

    parts, warnings = parse_sheet(path)
    print(f"파싱 완료: {len(parts)}개 품목")

    by_cat: dict[str, int] = {}
    for p in parts:
        by_cat[p["category"]] = by_cat.get(p["category"], 0) + 1
    for cat, n in sorted(by_cat.items(), key=lambda kv: -kv[1]):
        print(f"  {cat:16s} {n:4d}")

    no_id = sum(1 for p in parts if p["part_no"] is None)
    sold_out = sum(1 for p in parts if p["sold_out"])
    print(f"\n고유번호 없음 {no_id}건 (자연키로 임포트) / 품절 표시 {sold_out}건")

    skipped = [w for w in warnings if w.startswith("매핑에 없는")]
    for w in skipped:
        print(f"  ⚠️  {w}")

    if args.dry_run:
        print("\n--dry-run: DB에 쓰지 않았습니다. 샘플 3건:")
        for p in parts[:3]:
            print(" ", json.dumps(p, ensure_ascii=False)[:200])
        return

    url, key = load_env()
    print(f"\nSupabase로 upsert 중 … ({url})")
    n = upsert(url, key, parts)
    print(f"완료: {n}건 반영")


if __name__ == "__main__":
    main()
